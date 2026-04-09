#!/usr/bin/env python3
"""
Fetch the latest ERS Livestock, Dairy & Poultry Outlook forecasts.
Scrapes https://www.ers.usda.gov/publications?series=LDPM to find the latest
red-meat-and-poultry-forecasts.xlsx download link.
"""
import json, os, re, sys, urllib.request, urllib.error
from datetime import datetime

HEADERS = {"User-Agent": "HowardsHeuristics/1.0", "Accept": "text/html,application/xhtml+xml,*/*"}
ERS_BASE = "https://www.ers.usda.gov"
SERIES_URL = f"{ERS_BASE}/publications?series=LDPM"

# Fallback: known recent URLs (most recent first)
FALLBACK_URLS = [
    "https://ers.usda.gov/sites/default/files/_laserfiche/outlooks/113957/red-meat-and-poultry-forecasts.xlsx?v=94479",
]


def fetch_page(url, timeout=30):
    """Fetch a URL and return text content."""
    try:
        req = urllib.request.Request(url, headers=HEADERS)
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return resp.read().decode("utf-8", errors="replace")
    except Exception as e:
        print(f"  Failed to fetch {url[:80]}: {e}")
        return None


def discover_latest_outlook_url():
    """Scrape ERS publications page to find latest LDPM report and its Excel download."""
    print(f"  Fetching series page: {SERIES_URL}")
    html = fetch_page(SERIES_URL)
    if not html:
        return None

    # Find report links — look for links to individual report pages
    # Pattern: /publications/pub-details/?pubid=NNNNN or /webdocs/outlooks/NNNNN/
    report_links = re.findall(
        r'href="([^"]*(?:pub-details|publications/pub-details)[^"]*pubid=\d+[^"]*)"',
        html, re.IGNORECASE
    )
    
    if not report_links:
        # Try broader pattern for any outlooks links
        report_links = re.findall(
            r'href="(/[^"]*outlooks?/[^"]*\d{5,}[^"]*)"',
            html, re.IGNORECASE
        )
    
    if not report_links:
        # Try even broader - any link with LDPM in it
        report_links = re.findall(r'href="([^"]*[Ll][Dd][Pp][Mm][^"]*)"', html)

    print(f"  Found {len(report_links)} report link candidates")
    for link in report_links[:5]:
        print(f"    {link[:100]}")

    # Try each report link to find the Excel download
    for link in report_links[:5]:
        full_url = link if link.startswith("http") else ERS_BASE + link
        print(f"\n  Checking report page: {full_url[:80]}")
        report_html = fetch_page(full_url)
        if not report_html:
            continue

        # Look for the forecasts Excel link
        excel_links = re.findall(
            r'href="([^"]*red-meat-and-poultry-forecasts\.xlsx[^"]*)"',
            report_html, re.IGNORECASE
        )
        
        if not excel_links:
            # Broader: any xlsx link with "forecast" in it
            excel_links = re.findall(
                r'href="([^"]*forecast[^"]*\.xlsx[^"]*)"',
                report_html, re.IGNORECASE
            )

        if excel_links:
            download_url = excel_links[0]
            if not download_url.startswith("http"):
                download_url = ERS_BASE + download_url
            print(f"  Found Excel download: {download_url}")
            return download_url

    print("  Could not find Excel download link")
    return None


def parse_livestock_outlook(wb_data):
    """Parse the red-meat-and-poultry-forecasts.xlsx file.
    Returns dict with quarterly production, exports, imports, per capita for each species.
    """
    import openpyxl
    from io import BytesIO
    wb = openpyxl.load_workbook(BytesIO(wb_data), read_only=True, data_only=True)
    
    ws = wb["RMPFORE"] if "RMPFORE" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    
    # Parse header: Row 1 has years, Row 2 has quarter labels
    year_row = rows[1]
    qtr_row = rows[2]
    
    # Build column map: [(year, quarter, col_index), ...]
    col_map = []
    current_year = None
    for ci in range(1, len(year_row)):
        yr = year_row[ci]
        if yr is not None:
            try:
                current_year = int(float(yr))
            except:
                pass
        qtr = str(qtr_row[ci]).strip() if ci < len(qtr_row) and qtr_row[ci] else ""
        if current_year and qtr:
            # Map Roman numerals to quarter names
            qtr_map = {"I": "Q1", "II": "Q2", "III": "Q3", "IV": "Q4", "Annual": "Annual"}
            q = qtr_map.get(qtr, qtr)
            if q.startswith("Annual"):
                q = "Annual"
            col_map.append((current_year, q, ci))
    
    print(f"  Column map: {len(col_map)} columns spanning {col_map[0][0]}-{col_map[-1][0]}")
    
    # Parse data rows by finding species in each section
    SPECIES_ROWS = {}  # {(section, species): row_index}
    section = None
    for i, row in enumerate(rows):
        if not row or not row[0]: continue
        raw_label = str(row[0])
        label = raw_label.strip()
        
        if label.startswith("Production"):
            section = "production"
        elif label.startswith("Per capita"):
            section = "per_capita"
        elif label.startswith("U.S. trade"):
            section = "trade"
        elif label.startswith("Market prices"):
            section = "prices"
        elif section and raw_label.startswith("   "):
            species = label.lower()
            # Normalize species names
            if species.startswith("beef") and section == "production":
                SPECIES_ROWS[("production", "beef")] = i
            elif species.startswith("pork") and section == "production":
                SPECIES_ROWS[("production", "pork")] = i
            elif species.startswith("broiler") and section == "production":
                SPECIES_ROWS[("production", "broiler")] = i
            elif species.startswith("turkey") and section == "production":
                SPECIES_ROWS[("production", "turkey")] = i
            elif species.startswith("beef") and section == "per_capita":
                SPECIES_ROWS[("per_capita", "beef")] = i
            elif species.startswith("pork") and section == "per_capita":
                SPECIES_ROWS[("per_capita", "pork")] = i
            elif species.startswith("broiler") and section == "per_capita":
                SPECIES_ROWS[("per_capita", "broiler")] = i
            elif species.startswith("turkey") and section == "per_capita":
                SPECIES_ROWS[("per_capita", "turkey")] = i
            elif "beef" in species and "export" in species and section == "trade":
                SPECIES_ROWS[("exports", "beef")] = i
            elif "beef" in species and "import" in species and section == "trade":
                SPECIES_ROWS[("imports", "beef")] = i
            elif "pork export" in species and section == "trade":
                SPECIES_ROWS[("exports", "pork")] = i
            elif "pork import" in species and section == "trade":
                SPECIES_ROWS[("imports", "pork")] = i
            elif "broiler" in species and "export" in species and section == "trade":
                SPECIES_ROWS[("exports", "broiler")] = i
            elif "turkey" in species and "export" in species and section == "trade":
                SPECIES_ROWS[("exports", "turkey")] = i
    
    print(f"  Found {len(SPECIES_ROWS)} species/section combos")
    
    def extract_values(row_idx):
        """Extract values for all columns from a row."""
        row = rows[row_idx]
        result = {}
        for year, qtr, ci in col_map:
            val = row[ci] if ci < len(row) else None
            if val is not None:
                try:
                    val = round(float(val), 1)
                except:
                    val = None
            key = (year, qtr)
            result[key] = val
        return result
    
    # Build output per species
    output = {}
    for species in ["beef", "pork", "broiler", "turkey"]:
        species_data = {}
        for field in ["production", "per_capita", "exports", "imports"]:
            key = (field, species)
            if key in SPECIES_ROWS:
                species_data[field] = extract_values(SPECIES_ROWS[key])
        
        if species_data:
            output[species] = species_data
            # Print summary
            years_covered = sorted(set(yr for yr, _ in list(species_data.values())[0].keys()))
            fields = list(species_data.keys())
            print(f"  {species}: {fields}, years {years_covered}")
    
    return output


def main():
    print("=" * 60)
    print("Livestock Outlook Fetch")
    print(f"Time: {datetime.now().isoformat()}")
    print("=" * 60)
    
    # Step 1: Discover latest URL
    print("\n-- Discovering latest outlook URL --")
    url = discover_latest_outlook_url()
    
    if not url:
        print("  Falling back to known URL")
        url = FALLBACK_URLS[0]
    
    # Step 2: Download
    print(f"\n-- Downloading: {url[:80]} --")
    try:
        req = urllib.request.Request(url, headers=HEADERS)
        with urllib.request.urlopen(req, timeout=60) as resp:
            data = resp.read()
        print(f"  OK: {len(data):,} bytes")
    except Exception as e:
        print(f"  Download failed: {e}")
        return 1
    
    # Step 3: Parse
    print("\n-- Parsing --")
    result = parse_livestock_outlook(data)
    
    # Step 4: Print results
    print(f"\n{'='*60}")
    print(f"RESULTS: {len(result)} species")
    for species, data in result.items():
        print(f"\n  {species}:")
        for field, values in data.items():
            # Show annual values
            annuals = {yr: v for (yr, q), v in values.items() if q == "Annual"}
            print(f"    {field}: {annuals}")
    
    return 0


if __name__ == "__main__":
    sys.exit(main())
