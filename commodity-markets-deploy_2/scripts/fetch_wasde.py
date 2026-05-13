#!/usr/bin/env python3
"""
WASDE Data Fetch v11
- Wheat Table06: col0=MY, col1=Type ("All wheat"), col2+=S&D data
- Oil crops: tries current URL, falls back to cached data
"""

import json, os, sys, time, urllib.request, urllib.error, zipfile
from io import BytesIO
from datetime import datetime, timedelta

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(SCRIPT_DIR)
OUTPUT_DIR = os.path.join(REPO_ROOT, "data")

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
    "Accept": "*/*",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "identity",
    "Referer": "https://www.usda.gov/oce/commodity-markets/wasde",
    "Connection": "keep-alive",
}

CORN_URL = "https://www.ers.usda.gov/media/5764/feed-grains-yearbook-tables-all-years.xlsx?v=77939"
OILCROPS_URLS = [
    "https://www.ers.usda.gov/media/5220/all-tables-oil-crops-yearbook.xlsx?v=57591",
    "https://www.ers.usda.gov/media/5220/all-tables-oil-crops-yearbook.xlsx",
    "https://www.ers.usda.gov/media/5219/all-tables-oil-crops-yearbook-complete-data-set-in-compressed-zip-file.zip?v=11593",
]
WHEAT_URL = "https://www.ers.usda.gov/media/5706/wheat-data-all-years.xlsx?v=53976"


def fetch_url(url, timeout=60, retries=3, delay=10):
    for attempt in range(1, retries + 1):
        print(f"  {'Try' if attempt==1 else 'Retry '+str(attempt)}: {url[:90]}...")
        try:
            req = urllib.request.Request(url, headers=HEADERS)
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                data = resp.read()
                print(f"  OK: {len(data):,} bytes")
                return data
        except urllib.error.HTTPError as e:
            print(f"  HTTP {e.code}")
            if e.code in (404, 410):
                return None  # Definitely not there — no retry
            # 403/5xx may be transient (CDN edge nodes / rate limits). Retry with backoff.
            if attempt < retries:
                time.sleep(delay)
        except Exception as e:
            print(f"  {e}")
            if attempt < retries:
                time.sleep(delay)
    return None

def fetch_with_fallbacks(urls, timeout=60):
    for url in urls:
        data = fetch_url(url, timeout)
        if data and len(data) > 100: return data
    return None

def to_float(v):
    if v is None: return None
    s = str(v).strip()
    if s in ("", "--", "NA", "N/A", "nan", "None"): return None
    try: return float(s.replace(",", ""))
    except ValueError: return None

def ri(v):
    return round(v) if v is not None else None

def r1(v):
    return round(v, 1) if v is not None else None

def pct(num, den):
    if num is not None and den is not None and den != 0:
        return round(num / den * 100, 1)
    return None


def parse_corn(wb_data):
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(wb_data), read_only=True, data_only=True)

    ws01 = wb["FGYearbookTable01"]
    rows01 = list(ws01.iter_rows(values_only=True))
    area_planted, area_harvested, yield_per_acre = {}, {}, {}
    in_corn = False
    for i in range(2, len(rows01)):
        row = rows01[i]
        if not row: continue
        commodity = str(row[0]).strip() if row[0] else ""
        if commodity == "Corn": in_corn = True
        elif commodity and in_corn: break
        if not in_corn: continue
        my = str(row[1]).strip() if row[1] else ""
        if "/" not in my or len(my) < 7: continue
        area_planted[my] = to_float(row[2])
        area_harvested[my] = to_float(row[3])
        yield_per_acre[my] = to_float(row[5])
    print(f"  Corn area: {len(area_planted)} years")

    ws04 = wb["FGYearbookTable04"]
    rows04 = list(ws04.iter_rows(values_only=True))
    header = rows04[1]
    ncols = len([c for c in header if c is not None])
    col_labels = [str(header[c]).strip().replace("\n", " ") if header[c] else f"Col{c}" for c in range(2, ncols)]

    current_my = None
    sd_years, sd_data = [], []
    for i in range(2, len(rows04)):
        row = rows04[i]
        if not row or len(row) < 2: continue
        col0 = str(row[0]).strip() if row[0] else ""
        if "/" in col0 and len(col0) >= 7: current_my = col0
        quarter = str(row[1]).strip() if row[1] else ""
        if quarter.startswith("MY") and current_my:
            sd_years.append(current_my)
            sd_data.append({label: to_float(row[2 + ci]) for ci, label in enumerate(col_labels)})
    print(f"  Corn S&D: {len(sd_years)} years")

    # --- Ethanol from Table31 ---
    # Table31: Corn FSI detail. Quarterly format like Table04.
    # Headers: Market year, Quarter, HFCS, Glucose/dextrose, Starch, Alcohol for fuel, Alcohol for beverages, Cereals/other, Seed, Total FSI
    ethanol_data = {}
    try:
        ws31 = wb["FGYearbookTable31"]
        rows31 = list(ws31.iter_rows(values_only=True))
        h31 = rows31[1]
        h31_labels = [str(h31[c]).strip().replace("\n", " ") if h31[c] else "" for c in range(2, len(h31))]
        print(f"  Table31 headers: {h31_labels}")
        # Find the "Alcohol for fuel" column index
        ethanol_col = None
        for ci, lbl in enumerate(h31_labels):
            if "alcohol for fuel" in lbl.lower() or "fuel" in lbl.lower():
                ethanol_col = ci
                break
        if ethanol_col is not None:
            print(f"  Ethanol column index: {ethanol_col}")
            current_my31 = None
            for i in range(2, len(rows31)):
                row = rows31[i]
                if not row or len(row) < 2: continue
                col0 = str(row[0]).strip() if row[0] else ""
                if "/" in col0 and len(col0) >= 7: current_my31 = col0
                quarter = str(row[1]).strip() if row[1] else ""
                if quarter.startswith("MY") and current_my31:
                    ethanol_data[current_my31] = to_float(row[2 + ethanol_col])
            print(f"  Ethanol: {len(ethanol_data)} years")
        else:
            print("  Ethanol column not found in Table31")
    except Exception as e:
        print(f"  Table31 error: {e}")

    years = sd_years
    rows_out = [
        {"label": "Area planted", "values": [r1(area_planted.get(y)) for y in years]},
        {"label": "Area harvested", "values": [r1(area_harvested.get(y)) for y in years]},
        {"label": "Yield per harvested acre", "values": [r1(yield_per_acre.get(y)) for y in years]},
    ]

    # Helper to get S&D values by column prefix
    def get_sd(prefix):
        key = next((k for k in col_labels if k.lower().startswith(prefix.lower())), None)
        return [ri(sd_data[yi].get(key)) if key else None for yi in range(len(years))]

    # Supply
    rows_out.append({"label": "Beginning stocks", "values": get_sd("Beginning stocks")})
    rows_out.append({"label": "Production", "values": get_sd("Production")})
    rows_out.append({"label": "Imports", "values": get_sd("Imports")})
    rows_out.append({"label": "Total supply", "values": get_sd("Total supply"), "bold": True})

    # Use — reordered: Feed/residual, FSI (=food/alcohol/industrial + seed), Ethanol (indent), Seed (indent), Total domestic
    fsi_raw = get_sd("Food, alcohol, and industrial")
    seed_vals = get_sd("Seed use")
    # FSI = food/alcohol/industrial + seed
    fsi_vals = []
    for i in range(len(years)):
        f, s = fsi_raw[i], seed_vals[i]
        if f is not None and s is not None:
            fsi_vals.append(f + s)
        elif f is not None:
            fsi_vals.append(f)
        else:
            fsi_vals.append(None)

    rows_out.append({"label": "Feed and residual", "values": get_sd("Feed and residual")})
    rows_out.append({"label": "Food, seed & industrial", "values": fsi_vals})
    rows_out.append({"label": "Ethanol", "values": [ri(ethanol_data.get(y)) for y in years], "indent": True})
    rows_out.append({"label": "Seed", "values": seed_vals, "indent": True})
    rows_out.append({"label": "Total domestic use", "values": get_sd("Total domestic"), "bold": True})
    rows_out.append({"label": "Exports", "values": get_sd("Exports")})
    rows_out.append({"label": "Total usage", "values": get_sd("Total use"), "bold": True})

    # Ending stocks
    es_vals = get_sd("Ending stocks")
    tu_vals = get_sd("Total use")
    rows_out.append({"label": "Ending stocks", "values": es_vals, "bold": True})
    rows_out.append({"label": "Stocks/use (%)", "values": [pct(es_vals[i], tu_vals[i]) for i in range(len(years))], "bold": True, "pct": True})

    return {"id": "corn", "label": "Corn", "years": years,
            "sections": [{"header": "Supply and disappearance", "unit": "million bushels", "rows": rows_out}]}


def find_sheet(wb, *candidates):
    """Find a sheet by trying multiple name candidates."""
    for name in candidates:
        if name in wb.sheetnames:
            return wb[name]
    # Case-insensitive fallback
    for name in candidates:
        for sn in wb.sheetnames:
            if sn.lower() == name.lower():
                return wb[sn]
    return None


def parse_soybeans(wb):
    # Area: tab02 or Table02
    ws02 = find_sheet(wb, "tab02", "Table02", "Table 02")
    area_data = {}
    if not ws02:
        print("  Soy area sheet not found")
    else:
        rows02 = list(ws02.iter_rows(values_only=True))
        print(f"  Soy area sheet: {len(rows02)} rows")
        # Print first rows for debugging
        for i in range(min(6, len(rows02))):
            print(f"    Row {i}: {[str(c)[:25] if c else '' for c in (rows02[i] or [])][:7]}")
        
        for i in range(2, len(rows02)):
            row = rows02[i]
            if not row: continue
            year_val = str(row[0]).strip() if row[0] else ""
            
            # Try marketing year format first (1980/81)
            if "/" in year_val and len(year_val) >= 6:
                my = year_val
                planted = to_float(row[1])
                harvested = to_float(row[2])
                yld = to_float(row[3])
                # Check units: if planted > 500, it's in 1,000 acres, convert to million
                if planted and planted > 500:
                    planted = round(planted / 1000, 1)
                if harvested and harvested > 500:
                    harvested = round(harvested / 1000, 1)
                area_data[my] = {"planted": r1(planted), "harvested": r1(harvested), "yield": r1(yld)}
            else:
                # Try calendar year format (1980 -> 1980/81)
                try:
                    yr = int(year_val)
                    my = f"{yr}/{str(yr+1)[-2:]}"
                    planted = to_float(row[1])
                    harvested = to_float(row[2])
                    yld = to_float(row[3])
                    if planted and planted > 500:
                        planted = round(planted / 1000, 1)
                    if harvested and harvested > 500:
                        harvested = round(harvested / 1000, 1)
                    area_data[my] = {"planted": r1(planted), "harvested": r1(harvested), "yield": r1(yld)}
                except (ValueError, TypeError):
                    continue
    print(f"  Soy area: {len(area_data)} years")
    if area_data:
        s = list(area_data.keys())[-1]
        print(f"    Sample {s}: {area_data[s]}")

    # S&D: tab3 or Table03
    ws03 = find_sheet(wb, "tab3", "Table03", "Table 03", "tab03")
    if not ws03:
        print("  Soy S&D sheet not found")
        return None
    rows03 = list(ws03.iter_rows(values_only=True))
    print(f"  Soy S&D sheet: {len(rows03)} rows")
    # Print first data rows for format detection
    for i in range(min(8, len(rows03))):
        print(f"    Row {i}: {[str(c)[:20] if c else '' for c in (rows03[i] or [])][:11]}")

    # Find first data row (has marketing year in col 0)
    data_start = None
    for i in range(len(rows03)):
        row = rows03[i]
        if not row: continue
        val = str(row[0]).strip() if row[0] else ""
        if "/" in val and len(val) >= 6:
            try:
                int(val[:4])
                data_start = i
                break
            except ValueError:
                pass
    
    if data_start is None:
        print("  Could not find soy S&D data start")
        return None

    print(f"  Soy S&D data starts row {data_start}")
    if data_start < len(rows03):
        fr = rows03[data_start]
        print(f"    First data: {[str(c)[:15] if c else '' for c in (fr or [])][:12]}")

    years, sd = [], []
    for i in range(data_start, len(rows03)):
        row = rows03[i]
        if not row: continue
        my = str(row[0]).strip() if row[0] else ""
        if "/" not in my or len(my) < 6: continue
        try: int(my[:4])
        except ValueError: continue
        years.append(my)
        # Hardcoded: Table03 columns B through K (indices 1-10)
        # B=Beg stocks, C=Production, D=Imports, E=Total supply,
        # F=Crush, G=Seed use, H=Feed and residual, I=Exports, J=Total usage, K=Ending stocks
        sd.append([to_float(row[c]) if c < len(row) else None for c in range(1, 11)])

    print(f"  Soy S&D: {len(years)} years")

    # Explicit labels matching columns B through K (indices 1-10)
    sd_labels = [
        "Beginning stocks", "Production", "Imports", "Total supply",
        "Crush", "Seed use", "Feed and residual use", "Exports",
        "Total usage", "Ending stocks"
    ]

    rows_out = [
        {"label": "Area planted", "values": [area_data.get(y, {}).get("planted") for y in years]},
        {"label": "Area harvested", "values": [area_data.get(y, {}).get("harvested") for y in years]},
        {"label": "Yield per harvested acre", "values": [area_data.get(y, {}).get("yield") for y in years]},
    ]

    # Build S&D rows with proper spacing and calculated total domestic use
    # Supply section
    rows_out.append({"label": "Beginning stocks", "values": [ri(sd[yi][0]) for yi in range(len(years))]})
    rows_out.append({"label": "Production", "values": [ri(sd[yi][1]) for yi in range(len(years))]})
    rows_out.append({"label": "Imports", "values": [ri(sd[yi][2]) for yi in range(len(years))]})
    rows_out.append({"label": "Total supply", "values": [ri(sd[yi][3]) for yi in range(len(years))], "bold": True})

    # Use section
    crush_vals = [ri(sd[yi][4]) for yi in range(len(years))]
    seed_vals = [ri(sd[yi][5]) for yi in range(len(years))]
    feed_vals = [ri(sd[yi][6]) for yi in range(len(years))]
    # Calculate total domestic use = crush + seed + feed/residual
    dom_vals = []
    for yi in range(len(years)):
        c, s, f = crush_vals[yi], seed_vals[yi], feed_vals[yi]
        if c is not None and s is not None and f is not None:
            dom_vals.append(c + s + f)
        elif c is not None:
            dom_vals.append(c + (s or 0) + (f or 0))
        else:
            dom_vals.append(None)

    rows_out.append({"label": "Crush", "values": crush_vals})
    rows_out.append({"label": "Seed", "values": seed_vals})
    rows_out.append({"label": "Feed and residual", "values": feed_vals})
    rows_out.append({"label": "Total domestic use", "values": dom_vals, "bold": True})
    rows_out.append({"label": "Exports", "values": [ri(sd[yi][7]) for yi in range(len(years))]})
    rows_out.append({"label": "Total usage", "values": [ri(sd[yi][8]) for yi in range(len(years))], "bold": True})

    # Ending stocks + stocks/use
    es_vals = [ri(sd[yi][9]) for yi in range(len(years))]
    tu_vals = [ri(sd[yi][8]) for yi in range(len(years))]
    rows_out.append({"label": "Ending stocks", "values": es_vals, "bold": True})
    rows_out.append({"label": "Stocks/use (%)", "values": [pct(es_vals[i], tu_vals[i]) for i in range(len(years))], "bold": True, "pct": True})

    return {"id": "soybeans", "label": "Soybeans", "years": years,
            "sections": [{"header": "Supply and disappearance", "unit": "million bushels", "rows": rows_out}]}


def parse_soymeal(wb):
    ws = find_sheet(wb, "tab4", "Table04", "Table 04", "tab04")
    if not ws:
        print("  Soy meal sheet not found")
        return None
    rows = list(ws.iter_rows(values_only=True))
    print(f"  Soy meal sheet: {len(rows)} rows")
    for i in range(min(8, len(rows))):
        print(f"    Row {i}: {[str(c)[:20] if c else '' for c in (rows[i] or [])][:11]}")

    # Find first data row
    data_start = None
    for i in range(len(rows)):
        row = rows[i]
        if not row: continue
        val = str(row[0]).strip() if row[0] else ""
        if "/" in val and len(val) >= 6:
            try:
                int(val[:4])
                data_start = i
                break
            except ValueError: pass
    if data_start is None:
        print("  Could not find meal data start")
        return None

    print(f"  Meal data starts row {data_start}")
    fr = rows[data_start]
    print(f"    First: {[str(c)[:15] if c else '' for c in (fr or [])][:11]}")

    years, sd = [], []
    for i in range(data_start, len(rows)):
        row = rows[i]
        if not row: continue
        my = str(row[0]).strip() if row[0] else ""
        if "/" not in my or len(my) < 6: continue
        try: int(my[:4])
        except ValueError: continue
        years.append(my)
        # Read columns B through J (indices 1-9)
        sd.append([to_float(row[c]) if c < len(row) else None for c in range(1, 10)])
    print(f"  Soy meal: {len(years)} years")

    # Table04 columns: B=Beg stocks, C=Production, D=Imports, E=Total supply,
    # F=Domestic use, G=Exports, H=Total disappearance(usage), I=Ending stocks, J=Price
    rows_out = []
    rows_out.append({"label": "Beginning stocks", "values": [ri(sd[yi][0]) for yi in range(len(years))]})
    rows_out.append({"label": "Production", "values": [ri(sd[yi][1]) for yi in range(len(years))]})
    rows_out.append({"label": "Imports", "values": [ri(sd[yi][2]) for yi in range(len(years))]})
    rows_out.append({"label": "Total supply", "values": [ri(sd[yi][3]) for yi in range(len(years))], "bold": True})

    rows_out.append({"label": "Domestic use", "values": [ri(sd[yi][4]) for yi in range(len(years))]})
    rows_out.append({"label": "Exports", "values": [ri(sd[yi][5]) for yi in range(len(years))]})
    rows_out.append({"label": "Total usage", "values": [ri(sd[yi][6]) for yi in range(len(years))], "bold": True})

    es_vals = [ri(sd[yi][7]) for yi in range(len(years))]
    rows_out.append({"label": "Ending stocks", "values": es_vals, "bold": True})

    return {"id": "soybean_meal", "label": "Soybean Meal", "years": years,
            "sections": [{"header": "Supply and disappearance", "unit": "1,000 short tons", "rows": rows_out}]}


def parse_soyoil(wb):
    ws = find_sheet(wb, "tab5", "Table05", "Table 05", "tab05")
    if not ws:
        print("  Soy oil sheet not found")
        return None
    rows = list(ws.iter_rows(values_only=True))
    print(f"  Soy oil sheet: {len(rows)} rows")
    for i in range(min(8, len(rows))):
        print(f"    Row {i}: {[str(c)[:20] if c else '' for c in (rows[i] or [])][:11]}")

    # Find first data row
    data_start = None
    for i in range(len(rows)):
        row = rows[i]
        if not row: continue
        val = str(row[0]).strip() if row[0] else ""
        if "/" in val and len(val) >= 6:
            try:
                int(val[:4])
                data_start = i
                break
            except ValueError: pass
    if data_start is None:
        print("  Could not find oil data start")
        return None

    print(f"  Oil data starts row {data_start}")
    fr = rows[data_start]
    print(f"    First: {[str(c)[:15] if c else '' for c in (fr or [])][:11]}")

    years, sd = [], []
    for i in range(data_start, len(rows)):
        row = rows[i]
        if not row: continue
        my = str(row[0]).strip() if row[0] else ""
        if "/" not in my or len(my) < 6: continue
        try: int(my[:4])
        except ValueError: continue
        years.append(my)
        # Read columns B through K (indices 1-10)
        sd.append([to_float(row[c]) if c < len(row) else None for c in range(1, 11)])
    print(f"  Soy oil: {len(years)} years")

    # Table05 columns: B=Beg stocks, C=Production, D=Imports, E=Total supply,
    # F=Domestic total, G=Biofuel, H=Food/feed/other industrial, I=Exports, J=Total disappearance(usage), K=Ending stocks
    rows_out = []
    rows_out.append({"label": "Beginning stocks", "values": [ri(sd[yi][0]) for yi in range(len(years))]})
    rows_out.append({"label": "Production", "values": [ri(sd[yi][1]) for yi in range(len(years))]})
    rows_out.append({"label": "Imports", "values": [ri(sd[yi][2]) for yi in range(len(years))]})
    rows_out.append({"label": "Total supply", "values": [ri(sd[yi][3]) for yi in range(len(years))], "bold": True})

    rows_out.append({"label": "Domestic total", "values": [ri(sd[yi][4]) for yi in range(len(years))]})
    rows_out.append({"label": "Biofuel", "values": [ri(sd[yi][5]) for yi in range(len(years))], "indent": True})
    rows_out.append({"label": "Food, feed, and other industrial", "values": [ri(sd[yi][6]) for yi in range(len(years))], "indent": True})
    rows_out.append({"label": "Exports", "values": [ri(sd[yi][7]) for yi in range(len(years))]})
    rows_out.append({"label": "Total usage", "values": [ri(sd[yi][8]) for yi in range(len(years))], "bold": True})

    es_vals = [ri(sd[yi][9]) for yi in range(len(years))]
    rows_out.append({"label": "Ending stocks", "values": es_vals, "bold": True})

    return {"id": "soybean_oil", "label": "Soybean Oil", "years": years,
            "sections": [{"header": "Supply and disappearance", "unit": "million pounds", "rows": rows_out}]}


def parse_wheat(wb_data):
    try:
        import pandas as pd
    except ImportError:
        print("  pandas not installed")
        return None

    xls = None
    for engine in ["calamine", "openpyxl", None]:
        try:
            xls = pd.ExcelFile(BytesIO(wb_data), engine=engine)
            print(f"  Wheat opened with: {engine}")
            break
        except: pass
    if xls is None: return None

    # Area from Table01: col0=Class, col1=MY, col2=Planted, col3=Harvested, col4=Production, col5=Yield
    # Filter for "All wheat"
    area_data = {}
    try:
        df = pd.read_excel(xls, sheet_name="Table01", header=None)
        rows = [tuple(r) for r in df.values.tolist()]
        print(f"  Wheat area: Table01, {len(rows)} rows")
        for i in range(2, len(rows)):
            row = rows[i]
            cls = str(row[0]).strip().lower() if pd.notna(row[0]) else ""
            if cls != "all wheat": continue
            my = str(row[1]).strip() if pd.notna(row[1]) else ""
            if "/" not in my or len(my) < 6: continue
            try: int(my[:4])
            except ValueError: continue
            area_data[my] = {
                "planted": r1(to_float(row[2])),
                "harvested": r1(to_float(row[3])),
                "yield": r1(to_float(row[5])),
            }
        print(f"    All wheat area: {len(area_data)} years")
        if area_data:
            s = list(area_data.keys())[-1]
            print(f"    Sample {s}: {area_data[s]}")
    except Exception as e:
        print(f"    Area error: {e}")

    # S&D from Table06: col0=MY, col1=Type ("All wheat"), col2+=S&D data
    try:
        df = pd.read_excel(xls, sheet_name="Table06", header=None)
        rows = [tuple(r) for r in df.values.tolist()]
        print(f"  Wheat S&D: Table06, {len(rows)} rows")

        # Header is row 1: col0=MY label, col1=Type label, col2+=S&D column names
        header = rows[1]
        col_labels = []
        for c in range(2, len(header)):
            val = str(header[c]).strip() if pd.notna(header[c]) else ""
            if val:
                col_labels.append(val)
            else:
                break
        print(f"    Columns ({len(col_labels)}): {col_labels}")

        # Data rows: col0=MY, col1=Type, col2+=values
        # Filter for rows where col1 = "All wheat"
        years, sd_data = [], []
        for i in range(2, len(rows)):
            row = rows[i]
            wheat_type = str(row[1]).strip().lower() if pd.notna(row[1]) else ""
            if wheat_type != "all wheat": continue
            my = str(row[0]).strip() if pd.notna(row[0]) else ""
            if "/" not in my or len(my) < 6: continue
            try: int(my[:4])
            except ValueError: continue
            years.append(my)
            sd_data.append([to_float(row[2 + c]) for c in range(len(col_labels))])

        print(f"    All wheat S&D: {len(years)} years ({years[0] if years else '?'}..{years[-1] if years else '?'})")

        if not years:
            print("    No All wheat S&D found")
            return None

        # Build output
        rows_out = []
        if area_data:
            rows_out.append({"label": "Area planted", "values": [area_data.get(y, {}).get("planted") for y in years]})
            rows_out.append({"label": "Area harvested", "values": [area_data.get(y, {}).get("harvested") for y in years]})
            rows_out.append({"label": "Yield per harvested acre", "values": [area_data.get(y, {}).get("yield") for y in years]})

        renames = {}
        for lbl in col_labels:
            lower = lbl.lower()
            if "total disappearance" in lower: renames[lbl] = "Total usage"
            elif "total supply" in lower: renames[lbl] = "Total supply"
            elif "total domestic" in lower: renames[lbl] = "Total domestic use"
            elif lower.strip() == "food use": renames[lbl] = "Food"
            elif "seed use" in lower: renames[lbl] = "Seed"
            elif "feed and residual use" in lower: renames[lbl] = "Feed and residual"

        for ci, lbl in enumerate(col_labels):
            display = renames.get(lbl, lbl.replace(" 2/", "").replace(" 3/", "").replace(" 4/", "").strip())
            values = [ri(sd_data[yi][ci]) for yi in range(len(years))]
            rd = {"label": display, "values": values}
            if "total" in display.lower() or "ending" in display.lower(): rd["bold"] = True
            rows_out.append(rd)

        es = next((r for r in rows_out if "ending" in r["label"].lower()), None)
        tu = next((r for r in rows_out if r["label"] == "Total usage"), None)
        if es and tu:
            rows_out.append({"label": "Stocks/use (%)", "values": [pct(es["values"][i], tu["values"][i]) for i in range(len(years))], "bold": True, "pct": True})

        sample_rows = rows_out[:5]
        for r in sample_rows:
            print(f"    {r['label']}: ...{r['values'][-3:]}")

        return {"id": "wheat", "label": "Wheat", "years": years,
                "sections": [{"header": "Supply and disappearance", "unit": "million bushels", "rows": rows_out}]}

    except Exception as e:
        print(f"  S&D error: {e}")
        import traceback; traceback.print_exc()
        return None


# ══════════════════════════════════════════════════════════════
# LIVESTOCK BALANCE SHEETS — ERS Yearbook Data
# ══════════════════════════════════════════════════════════════

# ERS Livestock & Meat Domestic Data (beef, pork)
# ERS Poultry Yearbook (broiler, turkey)
# Multiple URL fallbacks since ERS periodically changes paths
# ERS Meat Supply & Disappearance — single file covers beef, pork, broiler, turkey
MEAT_SDP_URL = "https://www.ers.usda.gov/media/5531/meat-supply-and-disappearance-tables-historical.xlsx?v=84647"

# ERS Livestock, Dairy & Poultry Outlook — quarterly forecasts
# Update this URL when a new outlook is published (~monthly)
# Find latest at: https://www.ers.usda.gov/publications?series=LDPM
LDPO_URL = "https://ers.usda.gov/sites/default/files/_laserfiche/outlooks/113957/red-meat-and-poultry-forecasts.xlsx?v=94479"


def parse_meat_sdp_sheet(ws, species_label, min_year=1990):
    """Parse an ERS Meat S&D sheet. Extracts annual + quarterly rows.
    Cols: 0=Year, 1=Qtr, 2=CommProd, 3=FarmProd, 4=TotalProd, 5=_,
          6=BegStocks, 7=Imports, 8=TotalSupply, 9=Exports, 10=EndStocks,
          11=TotalDisappearance, 12=Population, 13=PC_carcass, 14=PC_retail
    """
    rows_data = list(ws.iter_rows(values_only=True))
    print(f"  {species_label}: sheet '{ws.title}', {len(rows_data)} rows")

    def to_int(v):
        if v is None: return None
        try: return round(float(str(v).replace(',', '').strip()))
        except: return None

    def to_1d(v):
        if v is None: return None
        try: return round(float(str(v).replace(',', '').strip()), 1)
        except: return None

    def extract_row(row):
        return {
            "production": to_int(row[4]),
            "beg_stocks": to_int(row[6]),
            "imports": to_int(row[7]),
            "total_supply": to_int(row[8]),
            "exports": to_int(row[9]),
            "end_stocks": to_int(row[10]),
            "total_use": to_int(row[11]),
            "per_capita": to_1d(row[14]) if len(row) > 14 else None,
        }

    current_year = None
    annual_data = []
    # quarterly: {1: [...], 2: [...], 3: [...], 4: [...]}
    qtr_data = {1: [], 2: [], 3: [], 4: []}

    QTR_MAP = {"Q1": 1, "Q2": 2, "Q3": 3, "Q4": 4}

    for i in range(3, len(rows_data)):
        row = rows_data[i]
        if not row or len(row) < 12: continue

        yr_cell = row[0]
        if yr_cell is not None:
            try:
                current_year = int(float(str(yr_cell).strip()))
            except (ValueError, TypeError):
                if str(yr_cell).strip().startswith(("1/", "2/", "3/", "Source", "Date")):
                    break
                continue

        if current_year is None or current_year < min_year or current_year > 2030:
            continue

        qtr = str(row[1]).strip() if row[1] else ""

        if "Yr" in qtr:
            d = extract_row(row)
            d["year"] = current_year
            annual_data.append(d)
        else:
            # Check for quarterly
            qtr_key = qtr[:2]  # "Q1", "Q2", etc.
            if qtr_key in QTR_MAP:
                d = extract_row(row)
                d["year"] = current_year
                qtr_data[QTR_MAP[qtr_key]].append(d)

    if not annual_data:
        print(f"  {species_label}: no annual rows found")
        return None

    years = [str(d["year"]) for d in annual_data]
    print(f"  {species_label}: {len(years)} annual years ({years[0]}..{years[-1]}), Q1={len(qtr_data[1])}, Q2={len(qtr_data[2])}, Q3={len(qtr_data[3])}, Q4={len(qtr_data[4])}")

    def build_rows(data_list):
        def vals(key):
            return [d[key] for d in data_list]
        rows_out = [
            {"label": "Beginning stocks", "values": vals("beg_stocks")},
            {"label": "Production", "values": vals("production")},
            {"label": "Imports", "values": vals("imports")},
            {"label": "Total supply", "values": vals("total_supply"), "bold": True},
            {"label": "Exports", "values": vals("exports")},
            {"label": "Ending stocks", "values": vals("end_stocks"), "bold": True},
            {"label": "Total use", "values": vals("total_use"), "bold": True},
            {"label": "Per capita disappearance (retail lbs)", "values": vals("per_capita"), "bold": True},
        ]
        return [r for r in rows_out if any(v is not None for v in r["values"])]

    QTR_LABELS = {1: "Jan-Mar", 2: "Apr-Jun", 3: "Jul-Sep", 4: "Oct-Dec"}

    sections = [{"header": "Annual", "unit": "million lbs (carcass weight)", "rows": build_rows(annual_data)}]

    for q in [1, 2, 3, 4]:
        if qtr_data[q]:
            qyears = [str(d["year"]) for d in qtr_data[q]]
            sections.append({
                "header": QTR_LABELS[q],
                "unit": "million lbs (carcass weight)",
                "rows": build_rows(qtr_data[q]),
                "years": qyears,
            })

    return {
        "id": species_label.lower(),
        "label": species_label,
        "years": years,
        "sections": sections,
    }



def parse_livestock_outlook(wb_data):
    """Parse the red-meat-and-poultry-forecasts.xlsx file from LDPO.
    Returns {species: {field: {(year, quarter): value}}}."""
    import openpyxl
    from io import BytesIO
    wb = openpyxl.load_workbook(BytesIO(wb_data), read_only=True, data_only=True)
    ws = wb["RMPFORE"] if "RMPFORE" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    # Column map from header rows
    year_row, qtr_row = rows[1], rows[2]
    col_map = []
    current_year = None
    for ci in range(1, len(year_row)):
        yr = year_row[ci]
        if yr is not None:
            try: current_year = int(float(yr))
            except: pass
        qtr = str(qtr_row[ci]).strip() if ci < len(qtr_row) and qtr_row[ci] else ""
        if current_year and qtr:
            q = {"I": "Q1", "II": "Q2", "III": "Q3", "IV": "Q4"}.get(qtr, "Annual" if "annual" in qtr.lower() else qtr)
            col_map.append((current_year, q, ci))

    # Detect species rows by section
    SPECIES_ROWS = {}
    section = None
    for i, row in enumerate(rows):
        if not row or not row[0]: continue
        raw_label = str(row[0])
        label = raw_label.strip()
        if label.startswith("Production"):
            section = "production"
        elif label.startswith("Per capita"):
            section = "per_capita"
        elif label.startswith("U.S. trade"):
            section = "trade"
        elif label.startswith("Market prices"):
            section = "prices"
        elif section and raw_label.startswith("   "):
            sp = label.lower()
            if sp.startswith("beef") and section == "production": SPECIES_ROWS[("production", "beef")] = i
            elif sp.startswith("pork") and section == "production": SPECIES_ROWS[("production", "pork")] = i
            elif sp.startswith("broiler") and section == "production": SPECIES_ROWS[("production", "broiler")] = i
            elif sp.startswith("turkey") and section == "production": SPECIES_ROWS[("production", "turkey")] = i
            elif sp.startswith("beef") and section == "per_capita": SPECIES_ROWS[("per_capita", "beef")] = i
            elif sp.startswith("pork") and section == "per_capita": SPECIES_ROWS[("per_capita", "pork")] = i
            elif sp.startswith("broiler") and section == "per_capita": SPECIES_ROWS[("per_capita", "broiler")] = i
            elif sp.startswith("turkey") and section == "per_capita": SPECIES_ROWS[("per_capita", "turkey")] = i
            elif "beef" in sp and "export" in sp and section == "trade": SPECIES_ROWS[("exports", "beef")] = i
            elif "beef" in sp and "import" in sp and section == "trade": SPECIES_ROWS[("imports", "beef")] = i
            elif "pork export" in sp and section == "trade": SPECIES_ROWS[("exports", "pork")] = i
            elif "pork import" in sp and section == "trade": SPECIES_ROWS[("imports", "pork")] = i
            elif "broiler" in sp and "export" in sp and section == "trade": SPECIES_ROWS[("exports", "broiler")] = i
            elif "turkey" in sp and "export" in sp and section == "trade": SPECIES_ROWS[("exports", "turkey")] = i

    def extract_values(row_idx):
        row = rows[row_idx]
        result = {}
        for year, qtr, ci in col_map:
            val = row[ci] if ci < len(row) else None
            if val is not None:
                try: val = round(float(val), 1)
                except: val = None
            result[(year, qtr)] = val
        return result

    output = {}
    for species in ["beef", "pork", "broiler", "turkey"]:
        species_data = {}
        for field in ["production", "per_capita", "exports", "imports"]:
            key = (field, species)
            if key in SPECIES_ROWS:
                species_data[field] = extract_values(SPECIES_ROWS[key])
        if species_data:
            output[species] = species_data

    print(f"  LDPO: {len(output)} species parsed")
    return output


def merge_ldpo_into_livestock(result, ldpo_data):
    """Merge LDPO quarterly forecasts into existing livestock balance sheet data.
    Updates production, imports, exports, per capita for quarterly sections."""
    QTR_HEADER_MAP = {"Q1": "Jan-Mar", "Q2": "Apr-Jun", "Q3": "Jul-Sep", "Q4": "Oct-Dec"}
    FIELD_TO_LABEL = {
        "production": "Production",
        "imports": "Imports",
        "exports": "Exports",
        "per_capita": "Per capita disappearance (retail lbs)",
    }

    for species_id, fields in ldpo_data.items():
        if species_id not in result.get("us", {}):
            continue
        entry = result["us"][species_id]
        sections = entry.get("sections", [])

        for field, values in fields.items():
            row_label = FIELD_TO_LABEL.get(field)
            if not row_label:
                continue

            for (yr, qtr), val in values.items():
                if val is None:
                    continue
                yr_str = str(yr)

                if qtr == "Annual":
                    continue  # WASDE handles annual data; LDPO only updates quarterly
                header = QTR_HEADER_MAP.get(qtr)
                if not header:
                    continue
                target_sections = [s for s in sections if s.get("header") == header]

                for section in target_sections:
                    sec_years = section.get("years", entry.get("years", []))
                    if yr_str in sec_years:
                        yi = sec_years.index(yr_str)
                        for row in section.get("rows", []):
                            if row["label"] == row_label:
                                while len(row["values"]) <= yi:
                                    row["values"].append(None)
                                # Round to int for S&D items, 1 decimal for per capita
                                if field == "per_capita":
                                    row["values"][yi] = round(val, 1)
                                else:
                                    row["values"][yi] = round(val)
                                break
                    elif qtr != "Annual":
                        # Year not in section yet — append it
                        sec_years.append(yr_str)
                        if section.get("years"):
                            section["years"] = sec_years
                        yi = len(sec_years) - 1
                        for row in section.get("rows", []):
                            while len(row["values"]) < yi:
                                row["values"].append(None)
                            if row["label"] == row_label:
                                row["values"].append(round(val, 1) if field == "per_capita" else round(val))
                            else:
                                row["values"].append(None)

    # ── Compute beginning stocks and total supply for quarterly 2026 ──
    for species_id in ["beef", "pork", "broiler", "turkey"]:
        if species_id not in result.get("us", {}):
            continue
        entry = result["us"][species_id]
        sections = entry.get("sections", [])

        # Get annual beginning stocks for 2026 from the annual section
        annual_section = next((s for s in sections if s.get("header", "").lower() in ("annual", "supply and disappearance")), None)
        annual_years = entry.get("years", [])
        annual_beg_2026 = None
        if annual_section and "2026" in annual_years:
            yi_2026 = annual_years.index("2026")
            beg_row = next((r for r in annual_section.get("rows", []) if r["label"] == "Beginning stocks"), None)
            if beg_row and yi_2026 < len(beg_row["values"]):
                annual_beg_2026 = beg_row["values"][yi_2026]

        # Also try to get Q4 ending stocks from previous year as fallback
        q4_section = next((s for s in sections if s.get("header") == "Oct-Dec"), None)
        q4_end_stocks_prev = None
        if q4_section:
            q4_years = q4_section.get("years", entry.get("years", []))
            if "2025" in q4_years:
                yi_q4 = q4_years.index("2025")
                es_row = next((r for r in q4_section.get("rows", []) if r["label"] == "Ending stocks"), None)
                if es_row and yi_q4 < len(es_row["values"]):
                    q4_end_stocks_prev = es_row["values"][yi_q4]

        # Q1 beginning stocks = annual 2026 beginning stocks (or Q4 2025 ending stocks)
        q1_beg = annual_beg_2026 if annual_beg_2026 is not None else q4_end_stocks_prev

        # Walk through quarters, carrying ending stocks forward as next quarter's beginning
        QTR_ORDER = ["Jan-Mar", "Apr-Jun", "Jul-Sep", "Oct-Dec"]
        prev_end_stocks = q1_beg

        for qi, qtr_header in enumerate(QTR_ORDER):
            qtr_section = next((s for s in sections if s.get("header") == qtr_header), None)
            if not qtr_section:
                continue

            sec_years = qtr_section.get("years", entry.get("years", []))
            if "2026" not in sec_years:
                continue
            yi = sec_years.index("2026")

            # Set beginning stocks
            if prev_end_stocks is not None:
                beg_row = next((r for r in qtr_section.get("rows", []) if r["label"] == "Beginning stocks"), None)
                if beg_row:
                    while len(beg_row["values"]) <= yi:
                        beg_row["values"].append(None)
                    beg_row["values"][yi] = round(prev_end_stocks)

            # Compute total supply = beg + prod + imports
            beg_val = None
            prod_val = None
            imp_val = None
            for row in qtr_section.get("rows", []):
                if row["label"] == "Beginning stocks" and yi < len(row["values"]):
                    beg_val = row["values"][yi]
                elif row["label"] == "Production" and yi < len(row["values"]):
                    prod_val = row["values"][yi]
                elif row["label"] == "Imports" and yi < len(row["values"]):
                    imp_val = row["values"][yi]

            if beg_val is not None and prod_val is not None:
                total_supply = round(beg_val + prod_val + (imp_val or 0))
                ts_row = next((r for r in qtr_section.get("rows", []) if r["label"] == "Total supply"), None)
                if ts_row:
                    while len(ts_row["values"]) <= yi:
                        ts_row["values"].append(None)
                    ts_row["values"][yi] = total_supply

            # For now, ending stocks are unknown (need cold storage data)
            # Set prev_end_stocks to None so subsequent quarters show None for beg stocks
            # until cold storage is wired up
            es_row = next((r for r in qtr_section.get("rows", []) if r["label"] == "Ending stocks"), None)
            if es_row and yi < len(es_row["values"]) and es_row["values"][yi] is not None:
                prev_end_stocks = es_row["values"][yi]
            else:
                prev_end_stocks = None

        if q1_beg is not None:
            print(f"  {species_id}: Q1 2026 beg stocks = {q1_beg}, computed total supply for available quarters")

    print(f"  LDPO merge complete")


def fetch_livestock_data(fetch_url_fn, fetch_fallbacks_fn, existing):
    """Fetch and parse all livestock species from ERS Meat S&D file."""
    import openpyxl
    from io import BytesIO
    results = {}

    print("\n  Fetching ERS Meat Supply & Disappearance...")
    data = fetch_url_fn(MEAT_SDP_URL)
    
    if data:
        try:
            wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
            print(f"  Sheets: {wb.sheetnames[:12]}")

            sheet_map = {
                "beef":    ("WASDE_Beef-Full", "Beef"),
                "pork":    ("WASDE_Pork-Full", "Pork"),
                "broiler": ("WASDE_Broiler-Full", "Broiler"),
                "turkey":  ("WASDE_Turkey-Full", "Turkey"),
            }
            
            for species_id, (sheet_name, label) in sheet_map.items():
                if sheet_name in wb.sheetnames:
                    r = parse_meat_sdp_sheet(wb[sheet_name], label)
                    if r:
                        results[species_id] = r
                else:
                    print(f"  {label}: sheet '{sheet_name}' not found")

        except Exception as e:
            print(f"  Meat S&D parse error: {e}")
            import traceback; traceback.print_exc()
    else:
        print("  Meat S&D download failed")

    # Preserve existing data for any species that failed
    for species in ["beef", "pork", "broiler", "turkey"]:
        if species not in results and species in existing:
            results[species] = existing[species]
            print(f"  Preserved existing {species} data")

    return results




# ══════════════════════════════════════════════════════════════
# TIER 1: Monthly WASDE Report (USDA OCE)
# Released ~noon ET on the 9th-12th of each month
# Contains latest 2-3 MY estimates for all commodities
# ══════════════════════════════════════════════════════════════


def build_wasde_report_urls():
    """Generate URL candidates for the current and previous month's WASDE report."""
    now = datetime.now()
    urls = []
    for months_back in [0, 1]:
        m = now.month - months_back
        y = now.year
        if m < 1:
            m += 12
            y -= 1
        mm = f"{m:02d}"
        yy = f"{y % 100:02d}"
        yyyy = str(y)
        # Primary USDA host
        urls.append(f"https://www.usda.gov/oce/commodity/wasde/wasde{mm}{yy}.xls")
        # USDA mirror via downloads subdomain (sometimes works when www.usda.gov 403s)
        urls.append(f"https://downloads.usda.gov/oce/commodity/wasde/wasde{mm}{yy}.xls")
        # USDA Markets directory (newer naming used since 2024)
        urls.append(f"https://www.usda.gov/sites/default/files/documents/wasde{mm}{yy}.xls")
        urls.append(f"https://www.usda.gov/sites/default/files/wasde{mm}{yy}.xls")
        # Cornell mirror (historical fallback)
        urls.append(f"https://usda.library.cornell.edu/apod/wasde{mm}{yy}.xls")
    return urls



class XlrdWrapper:
    """Minimal wrapper around xlrd workbook to match openpyxl interface."""
    def __init__(self, xls_wb):
        self._wb = xls_wb
        self.sheetnames = xls_wb.sheet_names()
        self._sheets = {}
        for name in self.sheetnames:
            self._sheets[name] = XlrdSheetWrapper(xls_wb.sheet_by_name(name))
        self.worksheets = [self._sheets[n] for n in self.sheetnames]
    def __getitem__(self, name):
        return self._sheets[name]

class XlrdSheetWrapper:
    """Minimal wrapper around xlrd sheet."""
    def __init__(self, sheet):
        self._sheet = sheet
        self.title = sheet.name
    def iter_rows(self, values_only=False):
        for i in range(self._sheet.nrows):
            row = []
            for j in range(self._sheet.ncols):
                cell = self._sheet.cell(i, j)
                val = cell.value
                if val == '': val = None
                row.append(val)
            yield tuple(row)


def parse_wasde_report(wb_data):
    """Parse the monthly WASDE report .xls file.
    Returns dict: {commodity_id: {years: [...], rows: [...]}}
    """
    from io import BytesIO
    wb = None
    try:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(wb_data), read_only=True, data_only=True)
        print(f"  Opened with openpyxl")
    except Exception:
        pass
    if wb is None:
        try:
            import xlrd
            xls_wb = xlrd.open_workbook(file_contents=wb_data)
            wb = XlrdWrapper(xls_wb)
            print(f"  Opened with xlrd")
        except Exception as e:
            print(f"  Could not open WASDE report: {e}")
            return {}

    print(f"  WASDE report sheets: {wb.sheetnames[:25]}")
    results = {}

    # ── Find pages by scanning sheet content ──
    page_map = {}  # {purpose: sheet_name}
    for sn in wb.sheetnames:
        rows = list(wb[sn].iter_rows(values_only=True))
        text = ' '.join(str(c) for row in rows[:8] for c in row if c).lower()
        if 'u.s. feed grain and corn' in text:
            page_map['corn'] = sn
        elif 'u.s. soybeans and product' in text:
            page_map['soybeans'] = sn
        elif 'u.s. wheat supply and use' in text:
            page_map['wheat'] = sn
        elif 'u.s. meats supply and use' in text:
            page_map['meats'] = sn

    print(f"  Page map: {page_map}")

    # ── Parse grain pages (years in columns) ──
    for comm_id, page_key in [("corn", "corn"), ("soybeans", "soybeans"), ("wheat", "wheat")]:
        if page_key not in page_map:
            print(f"  {comm_id}: page not found")
            continue
        parsed = _parse_grain_page(wb[page_map[page_key]], comm_id, "CORN" if comm_id == "corn" else None)
        if parsed:
            results[comm_id] = parsed

    # ── Parse soymeal/soyoil from same soybeans page ──
    if "soybeans" in page_map:
        for sub_id, section_marker in [("soybean_meal", "SOYBEAN MEAL"), ("soybean_oil", "SOYBEAN OIL")]:
            parsed = _parse_grain_page(wb[page_map["soybeans"]], sub_id, section_marker)
            if parsed:
                results[sub_id] = parsed

    # ── Parse Page 32: U.S. Meats Supply and Use (years in rows) ──
    if "meats" in page_map:
        _parse_meats_page(wb[page_map["meats"]], results)

    return results


def _parse_grain_page(ws, comm_id, section_marker=None):
    """Parse a grain/oilseed page where years are in column headers."""
    all_rows = list(ws.iter_rows(values_only=True))

    # Find the section start if needed
    start_row = 0
    if section_marker:
        for i, row in enumerate(all_rows):
            if not row or not row[0]: continue
            label = str(row[0]).strip().upper()
            # Exact match or starts-with (avoids "U.S. Feed Grain and Corn..." title matching "CORN")
            if label == section_marker.upper() or label.startswith(section_marker.upper() + " "):
                start_row = i
                break
        if start_row == 0:
            print(f"  {comm_id}: section '{section_marker}' not found")
            return None

    # Find year header row — look for "2023/24" or "2024" pattern
    year_cols = {}
    header_row_idx = None
    for i in range(start_row, min(start_row + 15, len(all_rows))):
        row = all_rows[i]
        if not row: continue
        for ci, cell in enumerate(row):
            if cell is None: continue
            s = str(cell).strip()
            # Match "2023/24" or "2024/25 Est." or "2025/26 Proj."
            if '/' in s and len(s) >= 5:
                try:
                    yr = int(s[:4])
                    if 1990 <= yr <= 2030:
                        # Use the Apr column if duplicate years (take last occurrence)
                        my = s.split()[0]  # "2024/25"
                        if len(my) > 7:
                            my = my[:4] + '/' + my.split('/')[1][-2:]
                        year_cols[ci] = my
                except (ValueError, IndexError):
                    pass
            # Match "2024.0" or "2025 Est."
            elif s.replace('.0', '').strip().isdigit():
                try:
                    yr = int(float(s))
                    if 2020 <= yr <= 2030:
                        year_cols[ci] = str(yr)
                except: pass
        if len(year_cols) >= 2:
            header_row_idx = i
            break

    if not year_cols or header_row_idx is None:
        print(f"  {comm_id}: no year headers found")
        return None

    # Deduplicate: if same MY appears twice, keep the LAST column (Apr estimate)
    my_to_col = {}
    for ci in sorted(year_cols.keys()):
        my_to_col[year_cols[ci]] = ci
    sorted_mys = sorted(my_to_col.keys())
    sorted_cols = [my_to_col[my] for my in sorted_mys]
    years = sorted_mys

    print(f"  {comm_id}: years = {years} (cols {sorted_cols})")

    # Label matching
    LABEL_MAP = {
        "area planted": "Area planted",
        "area harvested": "Area harvested",
        "yield per harvested acre": "Yield per harvested acre",
        "beginning stocks": "Beginning stocks",
        "production": "Production",
        "imports": "Imports",
        "supply, total": "Total supply",
        "crushings": "Crush",
        "crush": "Crush",
        "feed and residual": "Feed and residual",
        "food, seed": "Food, seed & industrial",
        "food, seed & ind": "Food, seed & industrial",
        "ethanol": "Ethanol",
        "domestic, total": "Total domestic use",
        "domestic disappearance": "Domestic disappearance",
        "biofuel": "Biofuel",
        "food, feed": "Food, feed, other industrial",
        "exports": "Exports",
        "use, total": "Total use",
        "ending stocks": "Ending stocks",
        "ending stock": "Ending stocks",
        "seed": "Seed",
        "residual": "Residual",
        "food": "Food",
    }
    BOLD = {"Total supply", "Total domestic use", "Total use", "Ending stocks", "Domestic disappearance"}
    INDENT = {"Ethanol", "Biofuel", "Food, feed, other industrial", "Seed", "Residual", "Food"}

    rows_out = []
    # Scan for a "Total" or next section marker to know when to stop
    stop_marker = None
    if section_marker:
        stop_row = len(all_rows)
        for i in range(header_row_idx + 5, len(all_rows)):
            row = all_rows[i]
            if row and row[0] and str(row[0]).strip() == "Total":
                stop_row = i
                break
    else:
        stop_row = len(all_rows)

    # Debug: log all row labels seen on this page
    seen_labels = []
    unmatched_labels = []
    for i in range(header_row_idx + 1, stop_row):
        row = all_rows[i]
        if not row: continue
        label_raw = str(row[0]).strip() if row[0] else ""
        if not label_raw or label_raw == "Filler": continue
        label_lower = label_raw.lower().strip()
        label_upper = label_raw.upper().strip()

        # Stop at section boundaries
        if label_upper.startswith("SOYBEAN OIL") or label_upper.startswith("SOYBEAN MEAL"):
            if comm_id == "soybeans":
                break
        if label_upper == "TOTAL" and section_marker:
            break

        seen_labels.append(label_raw)
        matched_label = None
        for key, display in LABEL_MAP.items():
            if key in label_lower:
                matched_label = display
                break
        if not matched_label:
            unmatched_labels.append(label_raw)
            continue

        values = []
        for ci in sorted_cols:
            v = row[ci] if ci < len(row) else None
            if v is not None:
                try:
                    # Strip footnote markers (*, †, /N) and parenthetical notes before parsing
                    if isinstance(v, str):
                        v_clean = v.strip()
                        # Common WASDE footnote suffixes: "43.8 *", "1,234 1/", "5.2 (proj)"
                        import re as _re
                        v_clean = _re.sub(r"\s*[*†‡§¶]+\s*$", "", v_clean)  # asterisk-style
                        v_clean = _re.sub(r"\s+\d+/\s*$", "", v_clean)     # numeric footnote refs like " 4/"
                        v_clean = _re.sub(r"\s*\([^)]+\)\s*$", "", v_clean) # trailing parens
                        v_clean = v_clean.replace(",", "")
                        if v_clean.upper() in ("NA", "N/A", "-", "—", ""):
                            values.append(None)
                            continue
                        values.append(round(float(v_clean), 1))
                    else:
                        values.append(round(float(v), 1))
                except (ValueError, TypeError):
                    values.append(None)
            else:
                values.append(None)

        if any(v is not None for v in values):
            rd = {"label": matched_label, "values": values}
            if matched_label in BOLD: rd["bold"] = True
            if matched_label in INDENT: rd["indent"] = True
            if not any(r["label"] == matched_label for r in rows_out):
                rows_out.append(rd)

    # Debug: report what we saw
    if unmatched_labels:
        print(f"  {comm_id}: unmatched WASDE labels: {unmatched_labels[:10]}")
    if not rows_out:
        print(f"  {comm_id}: no data rows matched (saw {len(seen_labels)} labels)")
        return None

    # Stocks/use
    es = next((r for r in rows_out if r["label"] == "Ending stocks"), None)
    tu = next((r for r in rows_out if "Total use" in r["label"] or "Total" == r["label"]), None)
    if es and tu:
        su = []
        for i in range(len(years)):
            e = es["values"][i] if i < len(es["values"]) else None
            t = tu["values"][i] if i < len(tu["values"]) else None
            su.append(round(e / t * 100, 1) if e is not None and t and t != 0 else None)
        rows_out.append({"label": "Stocks/use (%)", "values": su, "bold": True, "pct": True})

    print(f"  {comm_id}: {len(rows_out)} rows, {len(years)} years")
    unit = "million bushels" if comm_id in ("corn", "soybeans", "wheat") else "million pounds" if "oil" in comm_id else "1,000 short tons"
    return {"years": years, "rows": rows_out, "unit": unit}


def _parse_meats_page(ws, results):
    """Parse Page 32: U.S. Meats Supply and Use.
    Layout: years in ROWS, species stacked vertically.
    Each species block: year | month | BegStocks | Prod | Imports | TotalSupply | Exports | EndStocks | TotalUse | PerCapita
    We extract the Apr column for 2024 actual, 2025 Est., 2026 Proj."""
    all_rows = list(ws.iter_rows(values_only=True))

    SPECIES_MAP = {
        "beef": "Beef",
        "pork": "Pork",
        "broiler": "Broiler",
        "turkey": "Turkey",
    }

    # Column indices from Page 32 layout
    # Col: 0=Item, 1=Year, 2=Month, 3=BegStocks, 4=Production, 5=Imports,
    #      6=TotalSupply, 7=Exports, 8=EndStocks, 9=TotalUse, 10=PerCapita

    for species_id, species_label in SPECIES_MAP.items():
        # Find the species header row
        species_row = None
        for i, row in enumerate(all_rows):
            if not row or not row[0]: continue
            cell = str(row[0]).strip().lower().replace('\n', ' ')
            if cell == species_label.lower() or cell.startswith(species_label.lower()):
                species_row = i
                break

        if species_row is None:
            print(f"  {species_id}: not found in meats page")
            continue

        # Extract 2024, 2025 Apr, 2026 Apr rows
        years_data = {}
        current_yr = None
        for i in range(species_row, min(species_row + 8, len(all_rows))):
            row = all_rows[i]
            if not row: continue
            # Check for Filler row (section separator)
            if str(row[3] or "").strip() == "Filler": break

            # Track year — col 1 may be year (2024.0), label ("2025 Est."), or empty
            yr_val = row[1] if len(row) > 1 and row[1] else None
            month_val = str(row[2]).strip() if len(row) > 2 and row[2] else ""

            if yr_val is not None:
                try:
                    yr = int(float(str(yr_val).replace(" Est.", "").replace(" Proj.", "").strip()))
                    current_yr = yr
                except:
                    # "2025 Est." -> extract the year part
                    yr_str = str(yr_val).strip().split()[0]
                    try:
                        current_yr = int(float(yr_str))
                    except:
                        pass
            
            yr = current_yr

            # Accept any monthly estimate OR no-month (actuals). Later rows for same year overwrite —
            # WASDE lists multiple monthly estimates per year vertically; the LAST one is the freshest.
            ml = month_val.lower()
            has_month = any(m in ml for m in ("jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"))
            is_no_month = month_val == "" and yr is not None

            if (has_month or is_no_month) and yr and 2020 <= yr <= 2030:
                def to_f(v):
                    if v is None: return None
                    try:
                        if isinstance(v, str):
                            s = v.strip()
                            import re as _re
                            s = _re.sub(r"\s*[*†‡§¶]+\s*$", "", s)
                            s = _re.sub(r"\s+\d+/\s*$", "", s)
                            s = s.replace(",", "")
                            if s.upper() in ("NA", "N/A", "-", "—", ""): return None
                            return round(float(s), 1)
                        return round(float(v), 1)
                    except: return None

                new_row = {
                    "beg_stocks": to_f(row[3]) if len(row) > 3 else None,
                    "production": to_f(row[4]) if len(row) > 4 else None,
                    "imports": to_f(row[5]) if len(row) > 5 else None,
                    "total_supply": to_f(row[6]) if len(row) > 6 else None,
                    "exports": to_f(row[7]) if len(row) > 7 else None,
                    "end_stocks": to_f(row[8]) if len(row) > 8 else None,
                    "total_use": to_f(row[9]) if len(row) > 9 else None,
                    "per_capita": to_f(row[10]) if len(row) > 10 else None,
                }
                # If existing entry for this year has more non-null values than new, keep existing.
                # Otherwise prefer new (latest estimate).
                existing_entry = years_data.get(str(yr))
                if existing_entry is None:
                    years_data[str(yr)] = new_row
                else:
                    new_nonnull = sum(1 for v in new_row.values() if v is not None)
                    ex_nonnull = sum(1 for v in existing_entry.values() if v is not None)
                    if new_nonnull >= ex_nonnull:
                        years_data[str(yr)] = new_row

        if not years_data:
            print(f"  {species_id}: no data extracted from meats page")
            continue

        # Drop years where all metric values are None (e.g., 2027 column header present
        # but no actual data rows yet from WASDE)
        years_data = {y: d for y, d in years_data.items()
                      if any(v is not None for v in d.values())}
        if not years_data:
            print(f"  {species_id}: all years had empty data, skipping")
            continue

        years = sorted(years_data.keys())
        def vals(key):
            return [years_data[y].get(key) for y in years]

        rows_out = [
            {"label": "Beginning stocks", "values": vals("beg_stocks")},
            {"label": "Production", "values": vals("production")},
            {"label": "Imports", "values": vals("imports")},
            {"label": "Total supply", "values": vals("total_supply"), "bold": True},
            {"label": "Exports", "values": vals("exports")},
            {"label": "Ending stocks", "values": vals("end_stocks"), "bold": True},
            {"label": "Total use", "values": vals("total_use"), "bold": True},
            {"label": "Per capita disappearance (retail lbs)", "values": vals("per_capita"), "bold": True},
        ]
        rows_out = [r for r in rows_out if any(v is not None for v in r["values"])]

        results[species_id] = {
            "years": years,
            "rows": rows_out,
            "unit": "million lbs",
        }
        print(f"  {species_id}: {len(years)} years {years}, {len(rows_out)} rows")


def merge_wasde_into_existing(existing_commodity, wasde_data, comm_id):
    """Merge WASDE report data (2-3 years) into existing yearbook data (full history).
    WASDE overwrites matching years; new years get appended."""
    if not existing_commodity or not wasde_data:
        return existing_commodity
    
    ex_years = existing_commodity.get("years", [])
    ex_sections = existing_commodity.get("sections", [])
    if not ex_years or not ex_sections:
        return existing_commodity
    
    wasde_years = wasde_data["years"]
    wasde_rows = wasde_data["rows"]
    # Normalize WASDE labels to match yearbook conventions
    WASDE_TO_YEARBOOK = {
        "Total use": "Total usage",
        "Domestic disappearance": "Domestic total",
        "Food, feed, other industrial": "Food, feed, and other industrial",
        "Residual": "Feed and residual",
    }
    # Also build reverse map for meal/oil where yearbook has different labels
    WASDE_TO_YEARBOOK_MEAL = {
        "Domestic disappearance": "Domestic use",
    }

    wasde_map = {}
    for wr in wasde_rows:
        wasde_map[wr["label"]] = wr
        # Also add normalized versions
        if wr["label"] in WASDE_TO_YEARBOOK:
            wasde_map[WASDE_TO_YEARBOOK[wr["label"]]] = wr
        if comm_id in ("soybean_meal",) and wr["label"] in WASDE_TO_YEARBOOK_MEAL:
            wasde_map[WASDE_TO_YEARBOOK_MEAL[wr["label"]]] = wr
    
    for wi, wy in enumerate(wasde_years):
        if wy in ex_years:
            yi = ex_years.index(wy)
            for section in ex_sections:
                # Only update annual section, not quarterly (WASDE doesn't have quarterly livestock)
                if section.get("header", "").lower() in ("jan-mar", "apr-jun", "jul-sep", "oct-dec"):
                    continue
                for row in section.get("rows", []):
                    if row["label"] in wasde_map:
                        wr = wasde_map[row["label"]]
                        if wi < len(wr["values"]) and wr["values"][wi] is not None:
                            while len(row["values"]) <= yi:
                                row["values"].append(None)
                            row["values"][yi] = wr["values"][wi]
        else:
            ex_years.append(wy)
            yi = len(ex_years) - 1
            for section in ex_sections:
                if section.get("header", "").lower() in ("jan-mar", "apr-jun", "jul-sep", "oct-dec"):
                    continue
                for row in section.get("rows", []):
                    wr = wasde_map.get(row["label"])
                    val = wr["values"][wi] if wr and wi < len(wr["values"]) else None
                    while len(row["values"]) < yi:
                        row["values"].append(None)
                    row["values"].append(val)
    
    existing_commodity["years"] = ex_years
    print(f"  {comm_id}: merged WASDE data, now {len(ex_years)} years")
    return existing_commodity



def main():
    import openpyxl
    
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_file = os.path.join(OUTPUT_DIR, "wasde.json")

    # Load existing data
    existing = {}
    if os.path.exists(output_file):
        try:
            with open(output_file) as f:
                existing = json.load(f)
            print(f"Loaded existing: us={list(existing.get('us', {}).keys())}")
        except: pass

    print("=" * 60)
    print("WASDE Data Fetch — Two-Tier Architecture")
    print(f"Time: {datetime.now().isoformat()}Z")
    print("=" * 60)

    result = {"fetched_at": datetime.now().isoformat() + "Z", "us": {}}
    
    # Start with existing data as base
    if "us" in existing:
        result["us"] = existing["us"].copy()
    
    # Preserve world data if present
    if "world" in existing:
        result["world"] = existing["world"]
    if "world_fetched_at" in existing:
        result["world_fetched_at"] = existing["world_fetched_at"]

    # Track commodities that already have current/future MY data — never overwrite them with yearbook
    # (yearbook lags by one MY for new crop years, so overwriting would lose 2026/27 etc.)
    # Scrub: drop year columns where every row has None across every section.
    # Prevents accumulating empty trailing columns (e.g., a 2027 column with no data).
    def _scrub_empty_year_columns(entry):
        if not isinstance(entry, dict): return
        years = entry.get("years", [])
        sections = entry.get("sections", [])
        if not years or not sections: return
        # For each year index, check if every row in every section is None.
        keep_mask = [False] * len(years)
        for section in sections:
            for row in section.get("rows", []):
                values = row.get("values", [])
                for yi, v in enumerate(values):
                    if yi < len(keep_mask) and v is not None:
                        keep_mask[yi] = True
        if all(keep_mask): return  # No empty columns
        dropped = [years[i] for i in range(len(years)) if not keep_mask[i]]
        # Filter years array
        entry["years"] = [years[i] for i in range(len(years)) if keep_mask[i]]
        # Filter all row values arrays
        for section in sections:
            for row in section.get("rows", []):
                vals = row.get("values", [])
                row["values"] = [vals[i] for i in range(len(vals)) if i < len(keep_mask) and keep_mask[i]]
        if dropped:
            print(f"  Scrubbed empty year columns: {dropped}")

    for cid, entry in result.get("us", {}).items():
        _scrub_empty_year_columns(entry)

    current_year = datetime.now().year
    result["_existing_protected"] = []
    for cid, entry in result.get("us", {}).items():
        years = entry.get("years", []) if isinstance(entry, dict) else []
        for y in years:
            # Match either "2026/27" or "2027" formats
            try:
                if "/" in str(y):
                    yr = int(str(y)[:4])
                else:
                    yr = int(str(y))
                if yr >= current_year:
                    result["_existing_protected"].append(cid)
                    break
            except (ValueError, TypeError):
                pass
    if result["_existing_protected"]:
        print(f"  Protected commodities (have ≥{current_year} data): {result['_existing_protected']}")

    # ══════════════════════════════════════════════════════════
    # TIER 1: Monthly WASDE Report (primary, time-sensitive)
    # ══════════════════════════════════════════════════════════
    print("\n" + "═" * 60)
    print("TIER 1: Monthly WASDE Report")
    print("═" * 60)
    
    wasde_urls = build_wasde_report_urls()
    wasde_data = None
    wasde_source = None
    
    for url in wasde_urls:
        # 3 retries with 15s backoff helps when CDN edge nodes are 403-throttling temporarily
        data = fetch_url(url, timeout=30, retries=3, delay=15)
        if data:
            wasde_data = data
            wasde_source = url
            break
    
    if wasde_data:
        print(f"\n  WASDE report downloaded: {len(wasde_data):,} bytes")
        print(f"  Source: {wasde_source}")
        try:
            wasde_parsed = parse_wasde_report(wasde_data)
            print(f"  Parsed {len(wasde_parsed)} commodities from WASDE report")
            
            for comm_id, wdata in wasde_parsed.items():
                if comm_id in result["us"]:
                    # Merge WASDE data into existing yearbook data
                    result["us"][comm_id] = merge_wasde_into_existing(
                        result["us"][comm_id], wdata, comm_id)
                else:
                    # No existing data — create new entry from WASDE alone
                    is_livestock = comm_id in ("beef", "pork", "broiler", "turkey")
                    label_map = {
                        "corn": "Corn", "soybeans": "Soybeans", "wheat": "Wheat",
                        "soybean_meal": "Soybean Meal", "soybean_oil": "Soybean Oil",
                        "beef": "Beef", "pork": "Pork", "broiler": "Broiler", "turkey": "Turkey",
                    }
                    result["us"][comm_id] = {
                        "id": comm_id,
                        "label": label_map.get(comm_id, comm_id.title()),
                        "years": wdata["years"],
                        "sections": [{"header": "Supply and disappearance",
                                      "unit": wdata["unit"],
                                      "rows": wdata["rows"]}],
                    }
                    print(f"  {comm_id}: created new entry ({len(wdata['years'])} years)")
            
            result["wasde_report_source"] = wasde_source
            result["wasde_report_fetched_at"] = datetime.now().isoformat() + "Z"
            result["_wasde_updated"] = list(wasde_parsed.keys())
            result["_wasde_livestock"] = {k: v for k, v in wasde_parsed.items() 
                                          if k in ("beef", "pork", "broiler", "turkey")}
        except Exception as e:
            print(f"  WASDE report parse error: {e}")
            import traceback; traceback.print_exc()
    else:
        print("  WASDE report: not available (all URLs failed)")
        print("  This is normal if the report hasn't been released yet today.")

    # ══════════════════════════════════════════════════════════
    # TIER 2: ERS Yearbook Data (historical back-series)
    # ══════════════════════════════════════════════════════════
    print("\n" + "═" * 60)
    print("TIER 2: ERS Yearbook Data (historical)")
    print("═" * 60)
    
    print("\n-- CORN --")
    data = fetch_url(CORN_URL)
    if data:
        r = parse_corn(data)
        if r:
            protected = result.get("_wasde_updated", []) + result.get("_existing_protected", [])
            if "corn" in protected:
                # WASDE-updated or has future-year data — merge yearbook as base, preserving newer values
                _merge_yearbook_base(result["us"]["corn"], r)
            else:
                result["us"]["corn"] = r

    print("\n-- OIL CROPS --")
    data = fetch_with_fallbacks(OILCROPS_URLS)
    if data:
        try:
            wb = None
            try:
                wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
                print(f"  Opened as xlsx, sheets: {wb.sheetnames[:10]}")
            except Exception:
                try:
                    zf = zipfile.ZipFile(BytesIO(data))
                    if "Soy.xlsx" in zf.namelist():
                        wb = openpyxl.load_workbook(BytesIO(zf.read("Soy.xlsx")), read_only=True, data_only=True)
                        print(f"  Opened Soy.xlsx from zip, sheets: {wb.sheetnames[:10]}")
                except Exception as e2:
                    print(f"  Could not open as xlsx or zip: {e2}")

            if wb:
                for fn, parser in [("soybeans", parse_soybeans), ("soybean_meal", parse_soymeal), ("soybean_oil", parse_soyoil)]:
                    r = parser(wb)
                    if r:
                        protected = result.get("_wasde_updated", []) + result.get("_existing_protected", [])
                        if fn in protected:
                            _merge_yearbook_base(result["us"][fn], r)
                        else:
                            result["us"][fn] = r
        except Exception as e:
            print(f"  Error: {e}")
    else:
        print("  All URLs failed — preserving existing soy data")

    print("\n-- WHEAT --")
    data = fetch_url(WHEAT_URL)
    if data:
        r = parse_wheat(data)
        if r:
            protected = result.get("_wasde_updated", []) + result.get("_existing_protected", [])
            if "wheat" in protected:
                _merge_yearbook_base(result["us"]["wheat"], r)
            else:
                result["us"]["wheat"] = r

    # ── LIVESTOCK (ERS yearbooks as fallback) ──
    print("\n-- LIVESTOCK (ERS yearbooks) --")
    existing_livestock = {}
    for sp in ["beef", "pork", "broiler", "turkey"]:
        if sp in existing.get("us", {}):
            existing_livestock[sp] = existing["us"][sp]
    livestock_results = fetch_livestock_data(fetch_url, fetch_with_fallbacks, existing_livestock)
    for sp, ldata in livestock_results.items():
        # Use fresh yearbook data for livestock UNLESS WASDE-updated or has future-year data
        protected = result.get("_wasde_updated", []) + result.get("_existing_protected", [])
        if sp in protected:
            # Merge yearbook as base, preserving newer (WASDE/existing) values
            _merge_yearbook_base(result["us"][sp], ldata)
        else:
            result["us"][sp] = ldata
        print(f"  {sp}: {len(result['us'].get(sp, {}).get('years', []))} years")
    if not livestock_results:
        print("  WARNING: No livestock data returned!")

    # ── Re-apply WASDE annual livestock data after yearbook rebuild ──
    wasde_livestock = result.pop("_wasde_livestock", {})
    for sp, wdata in wasde_livestock.items():
        if sp in result["us"]:
            merge_wasde_into_existing(result["us"][sp], wdata, sp)

    # ── LDPO Quarterly Forecasts ──
    print("\n-- LDPO QUARTERLY FORECASTS --")
    try:
        ldpo_raw = fetch_url(LDPO_URL)
        if ldpo_raw:
            ldpo_data = parse_livestock_outlook(ldpo_raw)
            if ldpo_data:
                merge_ldpo_into_livestock(result, ldpo_data)
        else:
            print("  LDPO download failed — quarterly forecasts not updated")
    except Exception as e:
        print(f"  LDPO error: {e}")

    # Preserve any commodities that failed entirely
    for cid in ["corn", "soybeans", "soybean_meal", "soybean_oil", "wheat", "beef", "pork", "broiler", "turkey"]:
        if cid not in result["us"] and cid in existing.get("us", {}):
            result["us"][cid] = existing["us"][cid]
            print(f"  Preserved existing {cid} data")

    # ══════════════════════════════════════════════════════════
    # RESULTS
    # ══════════════════════════════════════════════════════════
    print(f"\n{'='*60}")
    print(f"RESULTS: {len(result['us'])} commodities")
    for cid, d in result["us"].items():
        yrs = d.get("years", [])
        total = sum(len(s.get("rows", [])) for s in d.get("sections", []))
        print(f"  {cid}: {len(yrs)} years ({yrs[0] if yrs else '?'}..{yrs[-1] if yrs else '?'}), {total} rows")

    result.pop("_wasde_updated", None)
    result.pop("_existing_protected", None)
    result.pop("_wasde_livestock", None)
    print(f"\nWriting {output_file}")
    with open(output_file, "w") as f:
        json.dump(result, f)
    print(f"  {os.path.getsize(output_file):,} bytes")
    return 0


def _merge_yearbook_base(existing_entry, yearbook_entry):
    """Merge yearbook historical data as the base for an existing entry.
    Yearbook provides years that don't exist in existing; existing data takes precedence."""
    if not yearbook_entry or not existing_entry:
        return
    
    ex_years = existing_entry.get("years", [])
    yb_years = yearbook_entry.get("years", [])
    
    # Find yearbook years not in existing
    new_years = [y for y in yb_years if y not in ex_years]
    if not new_years:
        return
    
    # Build label -> values map from yearbook
    yb_rows_map = {}
    for section in yearbook_entry.get("sections", []):
        for row in section.get("rows", []):
            yb_rows_map[row["label"]] = {yb_years[i]: row["values"][i] 
                                          for i in range(min(len(yb_years), len(row["values"])))}
    
    # Prepend new years to existing
    combined_years = [y for y in yb_years if y not in ex_years] + ex_years
    combined_years_sorted = sorted(combined_years, key=lambda y: y.split('/')[0] if '/' in y else y)
    
    for section in existing_entry.get("sections", []):
        for row in section.get("rows", []):
            old_vals = {ex_years[i]: row["values"][i] for i in range(min(len(ex_years), len(row["values"])))}
            yb_vals = yb_rows_map.get(row["label"], {})
            # Build combined values: yearbook for old years, existing for recent
            new_vals = []
            for y in combined_years_sorted:
                if y in old_vals and old_vals[y] is not None:
                    new_vals.append(old_vals[y])
                elif y in yb_vals:
                    new_vals.append(yb_vals[y])
                else:
                    new_vals.append(None)
            row["values"] = new_vals
    
    existing_entry["years"] = combined_years_sorted


if __name__ == "__main__":
    sys.exit(main())
